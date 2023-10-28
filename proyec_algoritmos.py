import tkinter as tk
import openpyxl
from tkinter import ttk
from tkcalendar import Calendar
import os
from tkinter import messagebox
from openpyxl import workbook
def ocultar_ventanas():
    ventana_citas.withdraw()
    ventana_agendar.withdraw()
    ventana_editar_citas.withdraw()
    ventana_eliminar.withdraw()
    resultados_cita.delete(1.0, tk.END)
    resultados_text.delete(1.0, tk.END)

def mostrar_ventana_citas():
    ventana_citas.deiconify()

def mostrar_ventana_agendar():
    ventana_agendar.deiconify()

def mostrar_ventana_edicion():
    ventana_editar_citas.deiconify()

def mostrar_ventana_eliminar():
    ventana_eliminar.deiconify()

def toggle_calendario():
    if calendario.winfo_viewable():
        calendario.place_forget()
    else:
        calendario.place(x=250, y=50)

def guardar_cita():
    nombre_ = nombre.get()
    num_telefono = no_telefono.get()
    motivo_de_cita = motivo_cita.get()
    fecha_consulta = calendario.get_date()

    #campos obligatorios
    if nombre_ == "" or num_telefono == "" or motivo_de_cita == "" or fecha_consulta == "":
        messagebox.showerror("Error", "Rellene todos los campos porfavor")

    if not os.path.exists("archivos_excel_citas"):
        os.mkdir("archivos_excel_citas")

    fecha_consulta = calendario.get_date()
    fecha_consulta = fecha_consulta.replace("/", "_")  # reemplaza las barras inclinadas por guiones bajos
    nombre_archivo = f"archivos_excel_citas/{fecha_consulta}.xlsx"

    if os.path.exists(nombre_archivo): 
        messagebox.showerror("Error", "Esta fecha ya está agendada, pruebe con otra.")
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Nombre del paciente", "No. de telefono", "Motivo de la consulta", ])
        row = [nombre_, num_telefono, motivo_de_cita, fecha_consulta]
        sheet.append(row)

        workbook.save(nombre_archivo)
        messagebox.showinfo("Éxito", "Su cita ha sido guardada.")
def ver_citas_por_fecha ():
    fecha_seleccionada = fecha_buscar.get()
    fecha_seleccionada = fecha_seleccionada.replace("/", "_")
    nombre_archivo = f"archivos_excel_citas/{fecha_seleccionada}.xlsx"

    if not os.path.exists(nombre_archivo):
        resultados_text.delete(1.0, tk.END)  # Borrar el contenido actual
        resultados_text.insert(tk.END, "No hay citas programadas para esta fecha.")
        return

    workbook = openpyxl.load_workbook(nombre_archivo)
    sheet = workbook.active
    citas = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nombre_paciente, no_telefono, motivo_cita, fecha_consulta = row
        citas.append(f"Nombre: {nombre_paciente}, Teléfono: {no_telefono}, Motivo: {motivo_cita}")

    resultados_text.delete(1.0, tk.END)  # Borrar el contenido actual
    resultados_text.insert(tk.END, "\n".join(citas))
    
#ventanas
ventana_principal = tk.Tk()
ventana_principal.title("Citas Médicas Dr. Chapatín")
ventana_principal.geometry("600x600")

ventana_citas = tk.Toplevel()
ventana_citas.title("Registro de citas")
ventana_citas.geometry("600x600")
ventana_agendar = tk.Toplevel()
ventana_agendar.title("Agregar tu cita")
ventana_agendar.geometry("600x600")
ventana_editar_citas = tk.Toplevel()
ventana_editar_citas.title("Editar mi cita")
ventana_editar_citas.geometry("600x600")
ventana_eliminar = tk.Toplevel()
ventana_eliminar.title("Eliminar cita")
ventana_eliminar.geometry("600x600")

#Texto de ventanas
label1 = tk.Label(ventana_principal,text="¡Bienvenido!")
label1.pack(ipadx=36)
nombre_label= tk.Label(ventana_agendar, text="Nombre del Paciente: ")
nombre_label.grid(row=0, column=0)
telefono_label= tk.Label(ventana_agendar, text="No. de Telefono: ")
telefono_label.grid(row=1, column=0)
motivo_cita_label = tk.Label(ventana_agendar, text="Motivo de la cita: ")
motivo_cita_label.grid(row=2, column=0)
fecha_label= tk.Label(ventana_agendar, text="Fecha")
fecha_label.grid(row=3, column=0)

#entradas de texto
nombre = tk.Entry(ventana_agendar)
nombre.grid(row=0, column=1)
no_telefono = tk.Entry(ventana_agendar)
no_telefono.grid(row=1, column=1)
motivo_cita = tk.Entry(ventana_agendar)
motivo_cita.grid(row=2, column=1)

# calendario oculto
calendario = Calendar(ventana_agendar, selectmode="day", date_pattern="dd/mm/yyyy", locale="es_ES")
calendario.place_forget()

# Boton para mostrar/ocultar el calendario
icono_calendario = tk.Button(ventana_agendar, text="Calendario", compound="left", command=toggle_calendario)
icono_calendario.grid(row=3, column=1)

# Funcion para ver las citas de una fecha específica
def editar_citas():
    resultados_cita.delete(1.0, tk.END)
    fecha_selec = escribir_citas.get()
    fecha_selec = fecha_selec.replace("/", "_")
    data_change = escribir_cambio.get()
    data_change_symbol = data_change.replace("/","_")
    opciones_obt = menu_desplegable.get()
    nombre_archivo_editar = f"archivos_excel_citas/{fecha_selec}.xlsx"
    data_change_ruta = f"archivos_excel_citas/{data_change_symbol}.xlsx"


    if not os.path.exists(nombre_archivo_editar):
        resultados_cita.delete(1.0, tk.END)  # Borrar el contenido actual
        resultados_cita.insert(tk.END, "No hay citas programadas para esta fecha.")
        return
    libro = openpyxl.load_workbook(nombre_archivo_editar)
    hoja = libro.active
    if opciones_obt == "nombre del paciente":
        celda_nec = hoja['A2']
        celda_nec.value = data_change  
        resultados_cita.insert(tk.END, "Se realizó el cambio")
        libro.save(f"archivos_excel_citas/{fecha_selec}.xlsx")

    if opciones_obt == "número de teléfono":
        celda_num = hoja['B2']
        celda_num.value = data_change  
        resultados_cita.insert(tk.END, "Se realizó el cambio")
        libro.save(f"archivos_excel_citas/{fecha_selec}.xlsx")

    if opciones_obt == "motivo de consulta":
        celda_motivo = hoja['C2']
        celda_motivo.value = data_change  
        resultados_cita.insert(tk.END, "Se realizó el cambio")
        libro.save(f"archivos_excel_citas/{fecha_selec}.xlsx")
    if opciones_obt == "fecha de consulta":
        if fecha_selec == data_change_symbol or os.path.exists(data_change_ruta):
            resultados_cita.insert(tk.END, "Hay citas programadas para esta fecha.") 
        else:
            celda_fecha = hoja['D2']
            celda_fecha.value = data_change 
            libro.save(f"archivos_excel_citas/{data_change_symbol}.xlsx")
            os.remove(f"archivos_excel_citas/{fecha_selec}.xlsx")
            resultados_cita.insert(tk.END, "Se realizó el cambio")

    

#funcion para eliminar las citas

def elimina_cita ():
     change_space = fecha_citas_eliminar_entry.get()
     change_space = change_space.replace("/", "_")
     file_eliminate = f"archivos_excel_citas/{change_space}.xlsx"
     if os.path.exists(file_eliminate):
        os.remove(file_eliminate)
        messagebox.showinfo("Éxito", "Su cita ha sido eliminada.")
     else:
        messagebox.showerror("No existe esa cita")

boton_ver_citas = tk.Button(ventana_principal,text="Ver citas", command=mostrar_ventana_citas)
boton_ver_citas.pack(ipadx=36, ipady= 10)
boton_agendar_cita = tk.Button(ventana_principal,text="Agendar una cita", command=mostrar_ventana_agendar)
boton_agendar_cita.pack(ipadx=15, ipady= 10)
boton_editar_cita = tk.Button(ventana_principal,text="Editar una cita", command=mostrar_ventana_edicion)
boton_editar_cita.pack(ipadx=25, ipady= 10)
boton_eliminar_cita = tk.Button(ventana_principal, text="eliminar una cita", command=mostrar_ventana_eliminar)
boton_eliminar_cita.pack(ipadx=45, ipady=10)
boton_guardar_cita = tk.Button(ventana_agendar, text="Guardar cita", command=guardar_cita)
boton_guardar_cita.grid(ipadx=15, ipady=10)
boton_regresar_menu = tk.Button(ventana_agendar, text="Menu", command=ocultar_ventanas)
boton_regresar_menu.grid(ipadx=30, ipady=10, row=100, column=0)
boton_regresar_menu_dos = tk.Button(ventana_editar_citas, text="Menu", command=ocultar_ventanas)
boton_regresar_menu_dos.grid(ipadx=30, ipady=10, row=100, column=0)
boton_regresar_menu_tres = tk.Button(ventana_citas, text="Menu", command=ocultar_ventanas)
boton_regresar_menu_tres.grid(ipadx=30, ipady=10, row=100, column=5)
boto_regresar_menu_elimar = tk.Button(ventana_eliminar, text="Menu", command=ocultar_ventanas)
boto_regresar_menu_elimar.grid(row=7, column=0)

# Agregar un widget de texto para mostrar los resultados
resultados_text = tk.Text(ventana_citas, wrap=tk.WORD, height=10, width=40)
resultados_text.grid(row=5, columnspan=2, padx=10, pady=10)
resultados_cita = tk.Text(ventana_editar_citas, wrap=tk.WORD, height=10, width=40)
resultados_cita.grid(row=10, columnspan=2, padx=10, pady=10)

# Agregar una entrada de texto para ingresar la fecha
fecha_buscar_label = tk.Label(ventana_citas, text="Buscar citas para la fecha (dd/mm/yyyy):")
fecha_buscar_label.grid(row=3, column=0)
fecha_buscar = tk.Entry(ventana_citas)
fecha_buscar.grid(row=3, column=1)

# Botón para buscar citas por fecha
boton_buscar_citas_por_fecha = tk.Button(ventana_citas, text="Buscar citas por fecha", command=ver_citas_por_fecha)
boton_buscar_citas_por_fecha.grid(row=4, columnspan=2, ipadx=15, ipady=10)

#apartado para editar los textos
editar_citas_label = tk.Label(ventana_editar_citas, text="Ingrese la fecha de su cita (dia/mes/año): ")
editar_citas_label.grid(row=3, column=0)
escribir_cambio_label = tk.Label(ventana_editar_citas, text="Escriba el cambio:")
escribir_cambio_label.grid(row=5, column=0)
escribir_citas= tk.Entry(ventana_editar_citas)
escribir_citas.grid(row=3, column=1)
escribir_cambio = tk.Entry(ventana_editar_citas)
escribir_cambio.grid(row=5, column=1)
editar_citas_boton = tk.Button(ventana_editar_citas, text="Editar cita", command=editar_citas)
editar_citas_boton.grid(row=4, column=2)
opciones_seleccion = tk.StringVar()
options = ['nombre del paciente', 'número de teléfono', 'motivo de consulta','fecha de consulta']
label_desplegable = tk.Label(ventana_editar_citas, text="Selecciones una opción")
label_desplegable.grid(row=1, column=0)
menu_desplegable = ttk.Combobox(ventana_editar_citas, textvariable=opciones_seleccion, values=options)
menu_desplegable.grid(row=2, column=0)

#apartado para eliminar citas
fecha_citas_eliminar = tk.Label(ventana_eliminar, text="Escriba la fecha en formato dia/mes/año de la cita que desea eliminar: ")
fecha_citas_eliminar.grid(row=1, column=0)
fecha_citas_eliminar_boton = tk.Button(ventana_eliminar, text="Eliminar", command=elimina_cita)
fecha_citas_eliminar_boton.grid(row=1, column=3)
fecha_citas_eliminar_entry = tk.Entry(ventana_eliminar)
fecha_citas_eliminar_entry.grid(row=1, column=2)
entry_eliminar = fecha_citas_eliminar_entry.get()

ocultar_ventanas()
ventana_principal.mainloop()